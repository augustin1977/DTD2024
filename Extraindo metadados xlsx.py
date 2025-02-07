import os
import json
import math
from huggingface_hub import login
import openpyxl
from unidecode import unidecode
import pandas as pd
from nltk.corpus import stopwords
import nltk
import re
import matplotlib.pyplot as plt
import torch
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.feature_extraction.text import TfidfVectorizer
from transformers import AutoModel,AutoTokenizer
from datetime import date
from key import huggings_TOKEN

def normaliza_texto(texto):
	texto_normalizado = unidecode(texto)
	return texto_normalizado

#Peso para os anos mais recentes
def calcula_pontos(ano):
	ano_atual=date.today().year
	k=0.8 #5 anos para uma área/linha de pesquisa morrer
	result = (100*math.exp(k * (ano - ano_atual)))
	return result
def carrega_pesquisadores_ativos(endereco):
	workbook = openpyxl.load_workbook(endereco)
	sheet=workbook["Planilha1"]
	dados=[]
	for linha in range(2,sheet.max_row):
		dados.append((sheet["A"+str(linha)].value).upper())
	return dados
 
 
class Importacao:
	def __init__(self,pasta):
		self.pasta=pasta
		self.arquivos=os.listdir(pasta)
		self.dados=[]
		self.json={}
		self.stop_words=None
		
	def abre_arquivo(self,arquivo):
		if ".XLSX" in arquivo.upper():
			print(f"Processando o arquivo {arquivo}")
			self.workbook = openpyxl.load_workbook(arquivo)
			self.sheet=self.workbook["Sheet1"]
		else:
			print(f"Arquivo {arquivo} não é um arquivo .XLSX e não será importado")
			self.workbook=None
			self.sheet=None
		
	def importa_linha(self,linha):
		
		registro={}
		if self.sheet:
			registro['Tipo_relatorio']=self.sheet["A"+str(linha)].value
			registro['Numero_documento']=self.sheet["B"+str(linha)].value
			registro['Ano']=self.sheet["C"+str(linha)].value
			registro['Area']=self.sheet["D"+str(linha)].value
			registro['Laboratorio']=self.sheet["E"+str(linha)].value
			registro['CRD']=self.sheet["F"+str(linha)].value
			registro['Titulo']=self.sheet["J"+str(linha)].value.upper()
			registro['N_projeto']=self.sheet["K"+str(linha)].value
			registro['Data_emissão']=self.sheet["O"+str(linha)].value
			registro['Palavras_chave']=self.sheet["P"+str(linha)].value.upper().strip().strip(".").split(';')
			registro['Resumo']=self.sheet["Q"+str(linha)].value
			resumo_limpo= self.sheet["Q"+str(linha)].value.lower()
			resumo_limpo=re.sub(r'[^\w\s]', '', resumo_limpo) 
			#resumo_limpo = ' '.join([word for word in resumo_limpo.split() if word not in self.stop_words])
			resumo_limpo = " ".join(self.sheet["P"+str(linha)].value.lower().strip().strip(".").split(';'))+' '+resumo_limpo
			registro['Resumo_limpo']=resumo_limpo
			registro['Equipe']=[]
			equipe=self.sheet["S"+str(linha)].value.upper().replace("EQUIPE:","").replace("AUTOR:","").split(";")
			for pessoa in equipe:
				padrao = r'^(.*?)(?= –|-|/)'
				resultado = re.search(padrao, pessoa)
				if resultado:
					resultado=normaliza_texto(resultado.group(1).strip().replace("  "," ").upper())
				else:
					resultado=normaliza_texto(pessoa.upper().strip())
				registro['Equipe'].append(resultado)
				# ~ print(resultado)
				# ~ time.sleep(0.05)
		return registro
	def numero_linhas(self):
		if self.sheet:
			return self.sheet.max_row
		else:
			return 0
	def importa_dados(self):
		print("#####Importanto dados dos arquivos#####")
		nltk.download('stopwords')
		self.stop_words=set(stopwords.words('portuguese'))
		print(self.stop_words)
		for arquivo in self.arquivos:
			abre_arquivo=os.path.join(self.pasta,arquivo)
			self.abre_arquivo(abre_arquivo)
			for i in range(2,self.numero_linhas()):
				self.dados.append(self.importa_linha(i))
	def cria_json(self):
		print("##### Criando arquvo 'newData.json' #####")
		nome_arquivo="newData.json"
		for registro in self.dados:
			if not (registro['Numero_documento'] in self.json):
				self.json[registro['Numero_documento']]=registro
			else:
				self.json[registro['Numero_documento']+"A"]=registro
				print(f"Registro {registro['Numero_documento']} repetido, criando o registro {registro['Numero_documento']+'A'}")
		with open(os.path.join(self.pasta,nome_arquivo), 'w' ,encoding="utf-8") as file:
			file.write(json.dumps(self.json, sort_keys=True, indent=4,ensure_ascii=False))
	def recupera_json(self):
		for arquivo in self.arquivos:
			if ".json" in arquivo:
				arquivo_json=os.path.join(self.pasta,arquivo)
				with open(os.path.join(self.pasta,'newData.json'), 'r', encoding='utf-8') as f:
					dados = json.load(f)

					# Converta o JSON em um DataFrame
					df = pd.DataFrame.from_dict(dados, orient='index')
					df.reset_index(inplace=True)
					df.rename(columns={'index': 'Numero_documento'}, inplace=True)
					self.dataframe=df
	def get_df(self):
		print("Recuperando json")
		self.recupera_json()
		return self.dataframe
     
class AI:
	def __init__(self, df):
		print("#### Iniciando o treinamento ####")
		self.ai_model="neuralmind/bert-base-portuguese-cased"
		self.df = df
		self.tokenizer = AutoTokenizer.from_pretrained(self.ai_model)
		self.model = AutoModel.from_pretrained(self.ai_model)
		self.vectorizer = TfidfVectorizer()
		self.projetos_selecionados=[]
		if os.path.exists('tokens.json'):
			print("Carregando tokens existentes...")
			with open('tokens.json', 'r', encoding='utf-8') as f:
				tokens = json.load(f)
			self.df['Tokens'] = tokens
		else:
			print("Gerando tokens...")
			self.df['Tokens'] = self.df['Resumo_limpo'].apply(lambda x: self.tokenizer.tokenize(x))
			with open('tokens.json', 'w', encoding='utf-8') as f:
				json.dump(self.df['Tokens'].tolist(), f, ensure_ascii=False, indent=4)

		if os.path.exists('embeddings.npy'):
			print("Carregando embeddings existentes...")
			embeddings = np.load('embeddings.npy')
			self.df['Embedding'] = list(embeddings)
		else:
			print("Não encontrado arquivo de Embeddings")
			self.gerar_embeddings()
			embeddings = np.vstack(self.df['Embedding'].values)
			np.save('embeddings.npy', embeddings)

		self.tfidf_matrix = self.vectorizer.fit_transform(self.df['Resumo_limpo'])

	def gerar_embedding(self, texto):
		inputs = self.tokenizer(texto, return_tensors='pt', truncation=True, padding=True, max_length=512)
		with torch.no_grad():
			outputs = self.model(**inputs)
		embedding = outputs.last_hidden_state.mean(dim=1).squeeze().numpy()
		return embedding

	def gerar_embeddings(self):
		print("Gerando Embedings... e fazendo fine training")
		self.df['Embedding'] = self.df['Resumo_limpo'].apply(self.gerar_embedding)
		print("Embeddings gerados com sucesso!")

	def buscar_resposta(self, pergunta, top_n=3):
		print("Buscando projetos com essa temática....")
		# Gera o embedding da pergunta
		embedding_pergunta = self.gerar_embedding(pergunta)
		
		# Converte a coluna de embeddings para uma matriz numpy
		embeddings = np.vstack(self.df['Embedding'].values)
		
		# Calcula a similaridade entre a pergunta e todos os resumos
		similaridades = cosine_similarity([embedding_pergunta], embeddings)[0]
		
		# Encontra os índices dos top N projetos mais similares
		indices = np.argsort(similaridades)[-top_n:][::-1]  # Ordena do maior para o menor
		
		# Recupera os top N projetos
		projetos_selecionados = self.df.iloc[indices]
		
		# Soma as palavras-chave dos projetos selecionados
		pesquisadores_chaves = []
		for pesquisadores in projetos_selecionados['Equipe']:
			for pesquisador in pesquisadores:
				if pesquisador not in pesquisadores_chaves:
					pesquisadores_chaves.append(pesquisador)
		
		# Retorna os projetos selecionados e as palavras-chave somadas
		self.projetos_selecionados=projetos_selecionados
		return pesquisadores_chaves

	def recomendar_pesquisadores(self, lista_pesquisadores=[], pesquisadores_ativos=[],top_n=10):
		self.participantes = {}
		print("Buscando pesquisadores....")
		projetos_selecionados=list(self.projetos_selecionados["Titulo"])
		for _, row in self.df.iterrows():
			if row["Titulo"] in projetos_selecionados:
				for participante in row["Equipe"]:
						pontos=calcula_pontos(row["Ano"])
						if participante in lista_pesquisadores:
							if participante in self.participantes:
								self.participantes[participante] += pontos
							else:
								self.participantes[participante] = pontos
		self.pesquisadores_ordenados = sorted(self.participantes.items(), key=lambda x: x[1], reverse=True)
		print(self.pesquisadores_ordenados[:min(top_n*5,15)])
		lista_pesquisadores_recomendados=[]
		for pesquisador in self.pesquisadores_ordenados:
			if pesquisador[0] in pesquisadores_ativos:
				lista_pesquisadores_recomendados.append(pesquisador[0])

      
		return [pesquisador for pesquisador in lista_pesquisadores_recomendados[:top_n]]
################################################
def main():
	login(token=huggings_TOKEN)
	root=os.getcwd()
	
 
	endereco=os.path.join(root,"Extração de dados","Relatórios Biblioteca")
	print(endereco)
	planilha=Importacao(endereco)
	planilha.importa_dados()
	planilha.cria_json()
	df=planilha.get_df()
	print("Carregando Pesquisadores Ativos")
	arquivo_pesquisadores_ativos=os.path.join(root,"Extração de dados","pesquisadores_ativos","pesquisadores_ativos.xlsx")
	print(arquivo_pesquisadores_ativos)
	pesquisadores_ativos=carrega_pesquisadores_ativos(arquivo_pesquisadores_ativos)
	print(pesquisadores_ativos)
	ai=AI(df)

	print("\n")
	print("BEM VINDO AO SISTEMA DE CONSULTA DE PESQUISADORES")
	print("\n")

	while(1):
		NewSystem = input("Você gostaria de fazer uma consulta? (y ou n): ").lower()
		print("\n")
		if NewSystem == "y":
			pergunta=input("Fale sobre o tipo de problema que vc está precisando de ajuda: ")
			print("\n")
			resposta=ai.buscar_resposta(pergunta, top_n=10,)
			
			for _, projeto in ai.projetos_selecionados.iterrows():
				print(f"Título: {projeto['Titulo']}")
				# ~ print(f"Resumo: {projeto['Resumo']}")
				# ~ print(f"Palavras-chave: {projeto['Palavras_chave']}")
				# ~ print("-" * 50)
			
			pesquisadores_ordenados=ai.recomendar_pesquisadores(lista_pesquisadores=resposta,
                                                       top_n=5,
                                                       pesquisadores_ativos=pesquisadores_ativos)
			print()
			print("As pessoas que podem te ajudar são:",(pesquisadores_ordenados))
			print()
		else:
			print("Obrigado por utilizar nosso sistema")
			print("Desenvolvido por SmartConnect IPT")
			print('\n')
			break
if __name__ == "__main__":
    main()